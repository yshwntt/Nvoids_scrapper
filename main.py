from nvoid_scraper import create_driver, scrape_query, _pause
from job_parser import parse_job
from gmail_service import get_gmail_service, create_draft, get_gmail_profile
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
import time
import re
import html
import config

TRACKER_FILE = config.TRACKER_FILE
MAX_EMAILS_PER_RUN = config.MAX_EMAILS_PER_RUN
DELAY_BETWEEN_EMAILS = config.DELAY_BETWEEN_EMAILS
EMPLOYER_EMAIL = config.EMPLOYER_EMAIL
KEYWORDS = config.KEYWORDS


def extract_job_id(url):
    match = re.search(r"id=(\d+)", url or "")
    return match.group(1) if match else None


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9 ]+", " ", (text or "").lower()).strip()


def is_job_already_processed(job_id, title=None, description=None):
    """Check if job was already processed in previous runs."""
    if not os.path.exists(TRACKER_FILE):
        return False
    
    title_signature = normalize_text(title or "")
    desc_signature = normalize_text(description or "")
    try:
        wb = load_workbook(TRACKER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            existing_job_id = row[0]
            existing_title = row[1] or ""
            existing_title_signature = normalize_text(existing_title)
            existing_desc_signature = ""
            if len(row) > 7 and row[7]:
                existing_desc_signature = str(row[7])

            if job_id and existing_job_id == job_id:
                return True
            if title_signature and existing_title_signature == title_signature:
                return True
            if desc_signature and existing_desc_signature == desc_signature:
                return True
        return False
    except Exception as e:
        print(f"Error checking job history: {e}")
        return False


def is_title_relevant(title: str) -> bool:
    title = (title or "").lower()
    
    has_keyword = any(word in title for word in KEYWORDS)
    has_excluded = any(word in title for word in config.EXCLUDE_KEYWORDS)
    
    return has_keyword and not has_excluded


def is_relevant_job(title: str, description: str = "") -> bool:
    title = (title or "").lower()
    description = (description or "").lower()
    
    # Check if it matches keywords
    has_keyword = any(word in title for word in KEYWORDS)
    
    # Check if it has excluded keywords in title or description
    has_excluded = any(word in title for word in config.EXCLUDE_KEYWORDS) or \
                   any(word in description for word in config.EXCLUDE_KEYWORDS)
    
    return has_keyword and not has_excluded


def pick_best_email(emails):
    """Pick the best recruiter email, avoiding generic domains."""
    if not emails:
        return None
    
    # First, try to find a non-excluded domain email
    for email in emails:
        email_lower = email.lower()
        if not any(excluded in email_lower for excluded in config.EXCLUDE_DOMAINS):
            return email
    
    # If all are excluded domains, return None
    return None


def extract_recruiter_name(description):
    """Try to extract recruiter's first name from JD. Returns None if not found."""
    patterns = [
        r'From:\s*\n+\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s*[,\n]',
        r'(?:Regards|Thanks)[,\s]*\n+\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s*\n',
        r'(?:Hi|Hello)\s+(?:I am|My name is)\s+([A-Z][a-z]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, description or "")
        if match:
            name = match.group(1).strip().split()[0]  # use first name only
            if len(name) > 2:
                return name
    return None


def build_email_body(title, recruiter_email, description, posted_at=None):
    escaped_title = html.escape(title or "")

    escaped_email = html.escape(recruiter_email or "")
    # Strip everything from "Keywords" onwards
    desc = description or ""
    cut = re.search(r'\bkeywords\b', desc, re.IGNORECASE)
    if cut:
        desc = desc[:cut.start()].rstrip()
    escaped_desc = html.escape(desc)

    # Clean title for intro: cut at "Developer" or "Engineer" (inclusive)
    clean_title_match = re.search(r'(.*?\b(?:Developer|Engineer)\b)', title or "", re.IGNORECASE)
    intro_title = html.escape(clean_title_match.group(1).strip() if clean_title_match else (title or ""))

    recruiter_name = extract_recruiter_name(description)
    greeting = f"Hi {recruiter_name}," if recruiter_name else "Hi,"

    # Extract phone number from description if present
    phone_match = re.search(r'(\+?1?\s*[\(\-]?\d{3}[\)\-\s]\s*\d{3}[\-\s]\d{4})', description or "")
    phone_line = f"      <strong>Phone:</strong> {html.escape(phone_match.group(1).strip())}<br>\n" if phone_match else ""

    # Convert IST (UTC+5:30) to EST (UTC-5) = subtract 10h30m
    posted_est = (posted_at - timedelta(hours=10, minutes=30)) if posted_at else None

    # Cloud platform detection: AWS=1, Azure=2, GCP=3
    desc_lower = (description or "").lower()
    cloud_nums = []
    if "aws" in desc_lower or "amazon web services" in desc_lower:
        cloud_nums.append("1")
    if "azure" in desc_lower:
        cloud_nums.append("2")
    if "gcp" in desc_lower or "google cloud" in desc_lower:
        cloud_nums.append("3")
    cloud_suffix = " - " + ", ".join(cloud_nums) if cloud_nums else ""

    posted_line = f"      <strong>Posted:</strong> {posted_est.strftime('%I:%M %p, %d %b %Y')}{cloud_suffix}<br>\n" if posted_est else ""

    return f"""<html>
  <body style='font-family:Arial, sans-serif; font-size:14px; color:#111;'>
    <p>{greeting}</p>
    <p>This is Yashwanth Chowdary, a Senior {intro_title} with over a decade of experience across banking and healthcare domains, delivering scalable and cloud-based solutions.</p>
    <p>I am currently available for C2C opportunities and open to relocation. Please find my resume attached for your review.</p>
    <p>
      <strong>Visa:</strong> H1B<br>
      <strong>Total Experience:</strong> 10+
    </p>
    <hr>
    <h3 style='margin-bottom:4px;'>Job Details</h3>
    <p style='margin-top:0;'>
      <strong>Title:</strong> {escaped_title}<br>
      <strong>Recruiter Email:</strong> <span style='user-select:all;'>{escaped_email}</span><br>
{phone_line}{posted_line}    </p>
    <h4 style='margin-bottom:4px;'>Job Description</h4>
    <div style='white-space:pre-wrap; word-break:break-word; background:#f9f9f9; padding:12px; border:1px solid #ddd; border-radius:4px; font-family:Arial, sans-serif; font-size:14px;'>{escaped_desc}</div>
    <hr>
    <p>
      Best regards,<br>
      <strong>Yashwanth Chowdary</strong><br>
      &#128231; yshc42@gmail.com<br>
      &#128222; (603) 438-1895
    </p>
  </body>
</html>"""


def ensure_tracker_file():
    if not os.path.exists(TRACKER_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["job_id", "title", "link", "emails", "phones", "status", "description_signature", "timestamp"])
        wb.save(TRACKER_FILE)


def save_applied_job(job_id, title, link, emails, phones, status, description_signature=""):
    ensure_tracker_file()

    wb = load_workbook(TRACKER_FILE)
    ws = wb.active

    ws.append([
        job_id,
        title,
        link,
        ",".join(emails),
        ",".join(phones),
        status,
        description_signature,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    wb.save(TRACKER_FILE)


def _process_jobs(driver, jobs, service, seen_signatures, seen_description_signatures,
                  current_run_ids, drafted_count, skipped_duplicates, job_counter):
    """
    Process a batch of jobs. Returns (drafted_count, skipped_duplicates, job_counter, stop_all).
    stop_all=True means MAX_EMAILS_PER_RUN hit — caller should stop all queries.
    """
    stop_all = False
    for job in jobs:
        if drafted_count >= MAX_EMAILS_PER_RUN:
            stop_all = True
            break

        title = job.get("title", "")
        link = job.get("link", "")
        posted_at = job.get("posted_at")
        job_id = extract_job_id(link)
        signature = normalize_text(title)

        job_counter += 1
        print(f"\nProcessing job {job_counter}...")
        print("Title:", title)

        if is_job_already_processed(job_id, title):
            # If this job was drafted earlier in the current run (e.g. appeared in a
            # prior query), just skip it — don't stop, there may be newer jobs ahead.
            if job_id in current_run_ids or signature in current_run_ids:
                print("Skipped: duplicate from current run")
                skipped_duplicates += 1
                continue
            # Otherwise it's from a previous run — all subsequent jobs are old too.
            print("Skipped: already processed — done with this query (newer jobs come first)")
            break

        if not is_title_relevant(title):
            print("Skipped: not relevant by title")
            continue

        if signature and signature in seen_signatures:
            print("Skipped: duplicate title in current run")
            skipped_duplicates += 1
            continue

        seen_signatures.add(signature)

        try:
            parsed = parse_job(driver, job)
        except Exception as e:
            print(f"Error parsing job: {e}")
            continue

        emails = parsed.get("emails", [])
        phones = parsed.get("phones", [])
        desc = parsed.get("description", "")
        desc_signature = normalize_text(desc)

        if desc_signature and desc_signature in seen_description_signatures:
            print("Skipped: duplicate job description in current run")
            skipped_duplicates += 1
            continue

        if is_job_already_processed(job_id, title, desc):
            print("Skipped: duplicate (already processed by description)")
            skipped_duplicates += 1
            continue

        seen_description_signatures.add(desc_signature)

        print("Emails:", emails)

        if not emails:
            print("Skipped: no email")
            continue

        best_email = pick_best_email(emails)

        if not best_email:
            print("Skipped: no valid recruiter email")
            continue

        subject = f"Interested || {title[:80]}"
        body = build_email_body(title, best_email, desc, posted_at)
        try:
            draft = create_draft(service, best_email, subject, body, cc=EMPLOYER_EMAIL, html=True)
            print(f"Draft created for {best_email} (id={draft.get('id')})")
            save_applied_job(job_id, title, link, [best_email], phones, "draft_created", desc_signature)
            current_run_ids.add(job_id)
            current_run_ids.add(signature)
            drafted_count += 1
            time.sleep(DELAY_BETWEEN_EMAILS)
        except Exception as e:
            print(f"Draft creation failed: {e}")
            save_applied_job(job_id, title, link, emails, phones, "draft_failed")

    return drafted_count, skipped_duplicates, job_counter, stop_all


def main():
    try:
        driver, wait, cutoff = create_driver()
    except Exception as e:
        print(f"Error creating driver: {e}")
        return

    try:
        service = get_gmail_service()
        profile = get_gmail_profile(service)
        print(f"Authenticated Gmail: {profile.get('emailAddress')}")

        queries = config.SEARCH_QUERIES
        if isinstance(queries, str):
            queries = [queries]

        drafted_count = 0
        skipped_duplicates = 0
        job_counter = 0
        seen_signatures = set()
        seen_description_signatures = set()
        seen_urls = set()
        current_run_ids = set()

        for i, query in enumerate(queries):
            jobs = scrape_query(driver, wait, query, seen_urls, cutoff)
            drafted_count, skipped_duplicates, job_counter, stop = _process_jobs(
                driver, jobs, service,
                seen_signatures, seen_description_signatures,
                current_run_ids, drafted_count, skipped_duplicates, job_counter
            )
            if stop:
                print(f"Reached email limit — stopping all queries.")
                break
            if i < len(queries) - 1:
                _pause(2.0, 4.5)

        print(f"\nDone. Drafts created: {drafted_count}")
        if skipped_duplicates > 0:
            print(f"Duplicates skipped: {skipped_duplicates}")

    except Exception as e:
        print(f"Unexpected error in main: {e}")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()